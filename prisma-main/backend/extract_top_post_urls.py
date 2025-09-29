import openpyxl
import sys
import json

def extract_top_post_urls(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    # Find the sheet name case-insensitively (English or Spanish)
    sheet_name = None
    for name in wb.sheetnames:
        if name.strip().lower() in ["top posts", "publicaciones principales"]:
            sheet_name = name
            break
    if not sheet_name:
        raise ValueError("Sheet 'Top posts' or 'Publicaciones principales' not found in workbook.")
    ws = wb[sheet_name]

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 6:
        raise ValueError("Sheet does not have enough rows for top posts.")

    # Skip first two rows (comment + empty)
    header_row = rows[2]
    # Get up to 50 data rows (top 50 posts)
    data_rows = rows[3:53]  # Top 50 posts (rows 3 to 52 inclusive)

    # Assume left block starts at col 0, right block at col 5 (adjust if needed)
    left_offset = 0
    right_offset = 4

    # Extract header names
    left_headers = header_row[left_offset:right_offset]
    right_headers = header_row[right_offset:right_offset+5]

    # Mapping for Spanish headers to English keys
    header_map = {
        'url de la publicación': 'url',
        'post url': 'url',
        'fecha de publicación': 'publish_date',
        'post publish date': 'publish_date',
        'interacciones': 'engagements',
        'engagements': 'engagements',
        'impresiones': 'impressions',
        'impressions': 'impressions',
    }

    def extract_post(row, offset, headers, metric_name):
        post = {}
        for i, header in enumerate(headers):
            if header is None:
                continue
            h = header.strip().lower()
            key = header_map.get(h, None)
            if key == 'url':
                post['url'] = row[offset + i]
            elif key == 'publish_date':
                post['publish_date'] = row[offset + i]
            elif key == metric_name:
                post[metric_name] = row[offset + i]
        return post

    top_by_engagement = [extract_post(row, left_offset, left_headers, 'engagements') for row in data_rows]
    top_by_impressions = [extract_post(row, right_offset, right_headers, 'impressions') for row in data_rows]

    result = {
        "top_by_engagement": top_by_engagement,
        "top_by_impressions": top_by_impressions
    }
    return result

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python extract_top_post_urls.py <xlsx_path>")
        sys.exit(1)
    xlsx_path = sys.argv[1]
    result = extract_top_post_urls(xlsx_path)
    print(json.dumps(result, indent=2, ensure_ascii=False)) 